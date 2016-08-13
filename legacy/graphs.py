import openpyxl as pyxl
import matplotlib.pyplot as plt
from matplotlib import style
import numpy as np

#XLOC_000001


# a list of all the datasets.

GenderExp_book = 'Female_Male_exp_levels_norm.xlsx'
Immgen_book = 'ImmGen_sex_exp_levels_norm.xlsx'

# a style for the graphs
#style.use('fivethirtyeight')

# open the excel file with data, and go to the first sheet.
wb = pyxl.load_workbook(filename=Immgen_book, read_only=True)
sheet = wb.worksheets[0]

# create example graph.
# x axis will consist of the cell name, gender, and months.
# go from A2 to K2
rows_dict = {}

# get the information from the first line
head=[]
### rethink of extracting data, because it take a lot of time ###
for row in sheet.iter_rows('A1:BW1'):
	for cell in row:
		head.append(cell.value)
#rows_dict['head'] = head

head = head[5:]

orderd_cells = []
cells_lists = {}
for cell in head:
	name = cell.split('_')[0]
	if name in cells_lists:
		cells_lists[name].append(cell)
	else:
		cells_lists[name] = [cell]
		orderd_cells.append(name)

#for key in cells_lists:
	#print key
#print orderd_cells

"""
cells_lists['GN'] = head[5:11]
cells_lists['MF'] = head[11:17]
cells_lists['DC'] = head[17:23]
cells_lists['B1ab'] = head[23:29] # remember - B1ab = B1a
cells_lists['CD19'] = head[29:37] # note - special case. also CD19 = B
#cells_lists['NK'] = head[37:]
"""






#cells_lists['MF'] = head
#print cells_lists
#extract data and hash it
#change last i value to sheet.max_row
"""for i in range(63410,63411): # note that it takes 56.8 seconds to load 100 lines of the excel sheet to memory, so it will take 10 hours to load all the information to memory.
	row = []
	for j in range(1,sheet.max_column+1):
		row.append(sheet.cell(row=i,column=j).value)
	# create a new entry for the data, and set the key as the gene_name or the xloc index.
	if row[1] == '-':
		rows_dict[row[0]] = row
	else:
		rows_dict[row[1]] = row
"""
row = []
#for j in range(1,sheet.max_column+1):
#	row.append(sheet.cell(row=63410,column=j).value)
rows_dict['xist'] = row
#print row
print '1'
#print row
# extract the data required.
# the title of the graph will be the name of the gene.
# prepare data for plotting.
# change how cells_names are taken.
# for row do each cell graph

def autolabel(rects):
    # attach values to the bars
    # remember to fix trailing zero's
    for rect in rects:
        height = rect.get_height()
        ax.text(rect.get_x() + rect.get_width()/2., 1.01*height,
                '%f' % float(height),
                ha='center', va='bottom')

width = 0.35
#for gene in rows_dict:

i = 0
exp_level = rows_dict['xist'][5:]#current row
gene_name = str(rows_dict['xist'][1])
title = " ".join([gene_name,'exp level by time and gender'])
for cell_name in orderd_cells:
	fig, ax = plt.subplots()

	ax.set_title(title)

	cell_list = cells_lists[cell_name]
	
	number_of_cells = len(cell_list)	
	index = np.arange(number_of_cells/2)
	xindex = np.arange(number_of_cells)
	exp_level_female = exp_level[i:i+number_of_cells][::2]
	exp_level_male = exp_level[i:i+number_of_cells][1::2]

	print "index val = " + str(len(index))
	print "exp val = " + str(len(exp_level_female))
	print "exp val2 = "+ str(len(exp_level_male))
	rects_females = ax.bar(index,exp_level_female,width,color ='c')
	rects_males = ax.bar(index+width,exp_level_male,width,color = 'r')
	
	ax.set_ylabel('Exp level')
	#ax.set_xlabel('time?')
	ax.set_xticks((xindex+width)/2.0)
	ax.set_xticklabels(cell_list,rotation=30)
	plt.subplots_adjust(bottom=0.20)
	ax.legend((rects_males[0],rects_females[0]),('Males','Females'))

	autolabel(rects_males)
	autolabel(rects_females)
	print 'abcd'
	plt.show()

	i = i+number_of_cells

"""

cells_names = rows_dict['head'][5:19]
cells_names_males = cells_names[1::2]
cells_names_females = cells_names[::2]

gene_name = str(rows_dict['XLOC_000001'][1])

exp_level = rows_dict['XLOC_000001'][5:]
exp_level_male = exp_level[1::2]
exp_level_female = exp_level[::2]


rects_females = ax.bar(index,exp_level_female,width,color = 'c')
rects_males = ax.bar(index+width,exp_level_male,width,color='g')


plt.title = " ".join([gene_name,'exp level by time and gender'])
ax.set_ylabel('Exp level')
ax.set_xlabel('time')
ax.set_xticks((index+width)/2.0,cells_names)
ax.legend((rects_males[0],rects_females[0]),('Males','Females'))

def autolabel(rects):
    # attach values to the bars
    # remember to fix trailing zero's
    for rect in rects:
        height = rect.get_height()
        ax.text(rect.get_x() + rect.get_width()/2., 1.05*height,
                '%f' % float(height),
                ha='center', va='bottom')
autolabel(rects_males)
autolabel(rects_females)

plt.show()
# plot the graph.
# with style.
#fig = plt.figure()
#ax1 = fig.add_subplot(111)

plt.title = "Gene name: ".join(gene_name)
x_axis_nums = [i for i in range(len(cells_names))]
plt.xticks(x_axis_nums,cells_names,rotation=10)
plt.xlabel('Cells',rotation=10)
plt.ylabel('expression level /n Log(2)')


print len(x_axis_nums)
plt.bar(x_axis_nums,exp_level)
plt.show()
# problem - need to take apart the strings.
nums = [i for i in range(len(x_tics))]
print nums
print x_tics
print y_data
plt.xticks(nums,x_tics,rotation=20)
plt.subplots_adjust(bottom=0.20,left=0.10)
plt.bar(nums,y_data,width=0.5)
plt.xlabel('Cell',rotation=0)
plt.ylabel('Exp level',rotation=60)
#plt.plot(nums,y_data,color='red')
#plt.show()
"""


