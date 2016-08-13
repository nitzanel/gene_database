import openpyxl as pyxl
import matplotlib
# choose backend
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import style
import numpy as np

# data bases
GenderExp_book = 'Female_Male_exp_levels_norm.xlsx'
Immgen_book = 'ImmGen_sex_exp_levels_norm.xlsx'


# save to folder
folder_name = 'test_graphs'

wb = pyxl.load_workbook(filename=Immgen_book,read_only=True)
sheet = wb.worksheets[0]

rows_dict = {}
information = []

# iterate over the first row and store the info.
#for row in sheet.iter_rows('A1:BW1'):
#	for cell in row:
#		information.append(cell.value)

for col in range(1,sheet.max_column + 1):
	information.append(sheet.cell(row = 1,column=col).value)

# remove uneeded information
information = information[5:]

orderd_cells = []
cells_lists = {}

for cell in information:
	# get the name of each cell
	name = cell.split('_')[0]
	if name in cells_lists:
		cells_lists[name].append(cell)
	else:
		cells_lists[name] = [cell]
		orderd_cells.append(name)

# load data into the dictionary
start_row = 2
end_row = 2

ind = 63410 #xist row, check.
"""
for row in ws.get_squared_range(min_col=2, max_col=2, min_row=1, max_row=ws.max_row):
    for cell in row: # each row is always a sequence
"""
row = []
for i in range(start_row,end_row+1):
	row = []
	for j in range(1,sheet.max_column+1):
		row.append(sheet.cell(row=i,column=j).value)
	if row[1] == '-':
		# if there is no gene name, name it XLOC...
		rows_dict[row[0]] = row
	else:
		rows_dict[row[1]] = row

"""
Load some other information in.
"""
tmp_row = []
for row in sheet.get_squared_range(min_col=1,max_col=sheet.max_column+1,min_row=ind,max_row=ind):
	for cell in row:
		tmp_row.append(cell.value)
rows_dict[tmp_row[1]] = tmp_row

# something for later
def autolabel(rects):
    # attach values to the bars
    # TODO: fix trailing zero's
    for rect in rects:
        height = rect.get_height()
        ax.text(rect.get_x() + rect.get_width()/2. + 0.03,
        		1.0*height+0.02,
                '%f' % float(height),
                fontsize=9,fontweight='bold',
                ha='center', va='bottom')

bar_width = 0.35

fig, ax = plt.subplots()
ax.set_ylabel('Exp level')
ax.yaxis.set_visible(False)
plt.subplots_adjust(bottom=0.20)
ax.get_yaxis().set_ticks([])
ax.xaxis.set_ticks_position('bottom')
for pos in ['left','right','top']:
	ax.spines[pos].set_color('none')


for key in rows_dict:
	i = 0
	# key will be the current gene. starting from the sixth value.
	exp_level = rows_dict[key][5:]
	gene_name = str(key) 	#str(rows_dict[key][1])
	for cell_name in orderd_cells:
		graph_title = " ".join(['gene:',gene_name,'cell :',cell_name,'expression level by time and gender'])
		ax.set_title(graph_title,y=1.08)
		cell_list = cells_lists[cell_name]

		number_of_cells = len(cell_list)
		index = np.arange(number_of_cells/2)
		xindex = np.arange(number_of_cells)
		
		exp_level_female = exp_level[i:i+number_of_cells][::2]
		exp_level_male = exp_level[i:i+number_of_cells][1::2]

		rects_females = ax.bar(index,exp_level_female,bar_width,color='c')
		rects_males = ax.bar(index+bar_width,exp_level_male,bar_width,color='r')

		ax.set_xticks((xindex+bar_width)/2.0)
		ax.set_xticklabels(cell_list,rotation=50,fontsize=8)
		leg = ax.legend((rects_males[0],rects_females[0]),('Males','Females'),
						loc='upper right',
						bbox_to_anchor=(1.15,1.05),
						fancybox=True)
		leg.get_frame().set_alpha(0.3)

		autolabel(rects_males)
		autolabel(rects_females)

		savename = '_'.join([gene_name,cell_name])
		filedir = '/'.join([folder_name,savename])
		plt.savefig(filedir)
		# clear the axis information so it can be reused
		ax.cla()
		# increament the iterator to get the correct range of exp_levels.
		i = i+number_of_cells
plt.close(fig)
