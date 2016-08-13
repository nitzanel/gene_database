import openpyxl as pyxl
import numpy as np

class Loader():
	def __init__(self):
		self.worksheets = {}
		self.cell_ranges = {}
		self.indexed = False
		self.cellIndexed = False

	def getCellsRanges(self):
		if self.cellIndexed is not True:
			self.getSets()
			for data_set in self.data_sets:
				if data_set == 'blank':
					continue
				
				if data_set not in self.cell_ranges:
					self.cell_ranges[data_set] = {}

				self.loadSheet(self.data_sets[data_set])
				head = self.loadHead(self.data_sets[data_set])
				last_cell = ''
				last_index = 0
				for index,info in enumerate(head):
					if index < 5:
						continue
					cell = info.split('_')[0]
					if cell != last_cell:
						if cell not in self.cell_ranges[data_set]:
							self.cell_ranges[data_set][cell] = {}
							self.cell_ranges[data_set][cell]['start'] = index+1
							if last_cell != '':
								self.cell_ranges[data_set][last_cell]['end'] = index
						last_cell = cell
					last_index = index
				self.cell_ranges[data_set][last_cell]['end'] = index+1
		return self.cell_ranges				




	# get the data sets list.
	def getSets(self):
		# create a list of the data sets.
		data_sets = {}
		self.blank_set = {}
		data_sets['blank'] = 'Blank.xlsx'
		data_sets['GenderExp'] = 'Female_Male_exp_levels_norm.xlsx'
		data_sets['Immgen'] = 'ImmGen_sex_exp_levels_norm.xlsx'
		self.data_sets = data_sets
		return data_sets

	# takes a gene name, returns the row index it's in.
	def findRowMatch(self,gene_name):
		if self.indexed is not True:
			# call loadNames with a worksheet that is already loaded.
			self.getSets()
			self.loadSheet(self.data_sets['blank'])
			self.loadNames(self.worksheets[self.data_sets['blank']]) 
		if gene_name in self.rows_dict:
			#print 'found gene '
			return self.rows_dict[gene_name]
		#return (''.join(['error ',gene_name, ' doesnt exist in the data set']))
		#print 'BAD GENE NAME'
		return -1

	# takes a list of workbook_names, and load the sheets into a dictionary for further use if they are not there.
	def loadSheet(self,workbook_name): 
		if workbook_name not in self.worksheets:
			# load the sheet into memory
			print 'workbookname:'
			print workbook_name
			wb = pyxl.load_workbook(filename = workbook_name,read_only=True)
			self.worksheets[workbook_name] = wb.worksheets[0] # could use split to take the first part of the  _
		else:
			print 'alrdy loadeed'		

	# takes a sheet, loads the gene column into a dictionary.
	def loadNames(self,sheet):
		self.rows_dict = {}
		for index, row in enumerate(sheet.get_squared_range(min_col=2,max_col=2,min_row=2,max_row=sheet.max_row)):
			for cell in row:
				self.rows_dict[cell.value] = index + 2
		print 'called loadNames'
		self.indexed = True
		return self.rows_dict

	# loads the first row of the sheet into a list and returns it.
	def loadHead(self,sheet):
		info = []
		current_sheet = self.worksheets[sheet]
		for row in current_sheet.get_squared_range(min_col=1,max_col=current_sheet.max_column,min_row=1,max_row=1):
			for cell in row:
				info.append(cell.value)
		return info

	def loadRow(self,gene_name,sheet):
		row_index = self.findRowMatch(gene_name)
		row_data = []
		print 'this is the sheet'
		print sheet
		current_sheet = self.worksheets[sheet]
		for row in current_sheet.get_squared_range(min_col = 1,max_col = current_sheet.max_column,min_row=row_index,max_row=row_index):
			for cell in row:
				row_data.append(cell.value)
		return row_data

	def loadPartialRow(self,gene_name,cell_name,sheet,current_set):
		row_index = self.findRowMatch(gene_name)
		row_data = []
		current_sheet = self.worksheets[sheet]
		ranges = self.getCellsRanges()
		for row in current_sheet.get_squared_range(min_col=self.cell_ranges[current_set][cell_name]['start'],max_col = self.cell_ranges[current_set][cell_name]['end'],min_row=row_index,max_row=row_index):
			for cell in row:
				row_data.append(cell.value)
		return row_data

	# the function will return a dictionary of expression level of a specific gene in a specific cell in all the datasets.
	def loadCellSpecific(self,gene_name,cell_name):
		gene_cell_dict = {}
		sets_names = []
		if self.findRowMatch(gene_name) == -1:
			return -1

		for data_set in self.data_sets:
			if data_set == 'blank':
				continue
			self.loadSheet(self.data_sets[data_set])
		data = {}
		for data_set in self.data_sets:
			if data_set == 'blank':
				continue
			data[data_set] = self.loadPartialRow(gene_name,cell_name,self.data_sets[data_set],data_set)
		return data

	# takes a list of genes, loads them.
	# this will call the rest of the functions.
	# can improve reloading
	def loadGenes(self,genes_list,data_sets):
		data_sets_dict = {}
		for data_set in data_sets:
			print 'this is the data set'
			print data_set
			self.loadSheet(data_set) 
			genes_dict = {}
			genes_dict['head'] = self.loadHead(data_set)
			for gene_name in genes_list:
				genes_dict[gene_name] = self.loadRow(gene_name,data_set)
			data_sets_dict[data_set] = genes_dict
		return data_sets_dict

